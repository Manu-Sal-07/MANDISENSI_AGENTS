import re
import json
import urllib.request
import urllib.parse
from pathlib import Path
import openpyxl

root = Path('d:/BMS COLL/PROJECT/MS-AI')
xlsx = root / 'imag' / 'Literature_Survey.xlsx'
wb = openpyxl.load_workbook(xlsx)
entries = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if r[0] is None:
            continue
        first = str(r[0]).strip()
        if not first:
            continue
        try:
            float(first)
        except:
            continue
        title = str(r[1]).strip() if r[1] is not None else ''
        if not title or title.lower() in ('title', 'titlle'):
            continue
        pub = str(r[2]).strip() if r[2] is not None else ''
        abstract = str(r[3]).strip() if r[3] is not None else ''
        keywords = str(r[4]).strip() if r[4] is not None else ''
        year = r[6]
        if hasattr(year, 'year'):
            year = year.year
        year = str(year).strip() if year is not None else ''
        url = str(r[7]).strip() if r[7] is not None else ''
        entries.append({
            'title': title.strip("'\" "),
            'publication': pub.strip("'\" "),
            'abstract': abstract.strip("'\" "),
            'keywords': keywords.strip("'\" "),
            'year': year,
            'url': url.strip("'\" "),
        })


def get_json(url):
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=20) as r:
        return json.loads(r.read().decode('utf-8'))


def doi_from_url(u):
    if not u:
        return None
    m = re.search(r'/article/(10\.\d{4,9}/[^/\?]+)', u)
    if m:
        return m.group(1)
    m = re.search(r'doi\.org/(10\.\d{4,9}/[^/\?]+)', u)
    if m:
        return m.group(1)
    return None


def select_best(items, title):
    norm = lambda s: re.sub(r'[^0-9a-zA-Z]+', '', s.lower())
    t = norm(title)
    best = None
    best_score = -1
    for item in items:
        cand = ' '.join(item.get('title', []))
        c = norm(cand)
        score = 0
        if t and c:
            if c == t:
                score = 100
            else:
                words = re.findall(r'[a-z]{3,}', t)
                score = sum(1 for w in words if w in c)
        if score > best_score:
            best_score = score
            best = item
    return best if best_score >= 1 else None


def escape(v):
    v = str(v)
    v = v.replace('\\', '\\\\')
    for a, b in [('{', '\\{'), ('}', '\\}'), ('%', '\\%'), ('$', '\\$'), ('&', '\\&'), ('#', '\\#'), ('_', '\\_'), ('^', '\\^{}'), ('~', '\\~{}')]:
        v = v.replace(a, b)
    return v

used = set()
bib = []
summary = []
for e in entries:
    e['doi'] = ''
    e['authors'] = []
    e['journal'] = e['publication']
    e['crossref'] = False
    doi = doi_from_url(e['url'])
    if doi:
        try:
            data = get_json('https://api.crossref.org/works/' + urllib.parse.quote(doi))
            if 'message' in data:
                msg = data['message']
                e['doi'] = doi
                e['crossref'] = True
                if not e['journal'] and msg.get('container-title'):
                    e['journal'] = msg['container-title'][0]
                if not e['year']:
                    date = msg.get('published-print') or msg.get('published-online') or msg.get('issued')
                    if date and date.get('date-parts'):
                        e['year'] = str(date['date-parts'][0][0])
                if msg.get('author'):
                    e['authors'] = [f"{a.get('family','')} {a.get('given','')}".strip() for a in msg['author'] if a.get('family') or a.get('given')]
        except Exception:
            pass
    if not e['doi'] and e['title']:
        try:
            q = urllib.parse.quote(e['title'])
            search = get_json(f'https://api.crossref.org/works?query.title={q}&rows=3')
            if 'message' in search and search['message'].get('items'):
                best = select_best(search['message']['items'], e['title'])
                if best:
                    e['crossref'] = True
                    e['doi'] = best.get('DOI', '')
                    if not e['journal'] and best.get('container-title'):
                        e['journal'] = best['container-title'][0]
                    if not e['year']:
                        date = best.get('published-print') or best.get('published-online') or best.get('issued')
                        if date and date.get('date-parts'):
                            e['year'] = str(date['date-parts'][0][0])
                    if best.get('author'):
                        e['authors'] = [f"{a.get('family','')} {a.get('given','')}".strip() for a in best['author'] if a.get('family') or a.get('given')]
        except Exception:
            pass
    base = 'unk'
    if e['authors']:
        base = re.sub(r'[^A-Za-z0-9]+', '', e['authors'][0].split()[0].lower())
    else:
        words = re.findall(r'[A-Za-z]+', e['title'].lower())
        base = ''.join(w[0] for w in words[:2]) if words else 'p'
    year = e['year'] if e['year'] else '2025'
    tail = ''.join(w[0] for w in re.findall(r'[A-Za-z]+', e['title'].lower())[:2])
    key = re.sub(r'[^A-Za-z0-9]+', '', f'{base}{year}{tail}')
    if not key:
        key = 'paper' + year
    orig = key
    suffix = 1
    while key in used:
        key = f'{orig}{suffix}'
        suffix += 1
    used.add(key)
    auth = ' and '.join(e['authors']) if e['authors'] else 'Unknown'
    fields = {
        'title': e['title'],
        'author': auth,
        'journal': e['journal'],
        'year': year,
        'volume': '',
        'number': '',
        'pages': '',
        'doi': e['doi'],
        'url': e['url'],
        'abstract': e['abstract'],
        'keywords': e['keywords'],
    }
    bib.append(f'@article{{{key},')
    for f, v in fields.items():
        bib.append(f'  {f}={{{escape(v)}}},')
    bib[-1] = bib[-1].rstrip(',')
    bib.append('}\n')
    summary.append({'key': key, 'year': year, 'doi': bool(e['doi']), 'url': e['url']})

Path(root/'imag'/'references.bib').write_text('\n'.join(bib), encoding='utf-8')
Path(root/'imag'/'references_summary.json').write_text(json.dumps(summary, indent=2), encoding='utf-8')
print('wrote', len(summary), 'entries to imag/references.bib')
